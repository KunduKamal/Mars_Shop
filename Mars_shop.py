
import pandas as pd
import numpy as np 
import openpyxl 
from openpyxl import load_workbook

# Declear for global veriable 
monthly_fixed_cost = 800
int_rice_am = 300
int_lentil_am = 350
int_capital_am = 4000
rice_pp_cost = 5 
lentil_pp_cost = 7
profit_margin_rate = 0.20
monthly_sales_r = 0
monthly_sales_l = 0
stock_short_r = 100
stock_short_l = 100
result_r = []
result_l = [] 
week_cnt = 1
month_end_sales_r = []
month_end_sales_l = []

## ------------------------------------------------------------------------ ##
# import sales information (excel data) in two dimonesion array
file = pd.read_excel("Data/weekly_sales.xlsx", sheet_name="Sheet1", usecols="C,D")
a = dict(file)
sales_info = zip(list(a['Rice']), list( a['Lentil'] ) )
sales_info = list( sales_info )
#print( sales_info )

## ------------------------------------------------------------------------ ##
def purchase_cost(r1):
    if r1 == 'rice':
        c_cost = int_rice_am * (rice_pp_cost-(rice_pp_cost*profit_margin_rate))
    elif r1 == 'lentil':
        c_cost = int_lentil_am * (lentil_pp_cost-(lentil_pp_cost*profit_margin_rate))
    else:
        c_cost = 0
        print("Invalid input product")
    
    return c_cost

## ------------------------------------------------------------------------ ##
in_stock_r = int_rice_am
in_stock_l = int_lentil_am
monthly_p_cost = 0
month_end_cost = []
month_end_Tsales = []
## ------------------------------------------------------------------------ ##
for r in range(0, len(sales_info)):

    if in_stock_r  < stock_short_r:
        in_stock_r  += int_rice_am
        monthly_p_cost += purchase_cost('rice')
       
    if in_stock_l  < stock_short_l:
        in_stock_l  += int_lentil_am
        monthly_p_cost += purchase_cost('lentil')
    
    in_stock_r = in_stock_r - sales_info[r][0]
    in_stock_l = in_stock_l  - sales_info[r][1]
    
    monthly_sales_r += sales_info[r][0]
    monthly_sales_l += sales_info[r][1]
    

    result_r.append(in_stock_r)
    result_l.append(in_stock_l)
    
    if week_cnt < 4: 
        week_cnt += 1
    else:            
        week_cnt = 1
        #month_end_sales_r.append(monthly_sales_r * rice_pp_cost)
        #month_end_sales_l.append(monthly_sales_l * lentil_pp_cost)
        month_end_Tsales.append((monthly_sales_r * rice_pp_cost)+(monthly_sales_l * lentil_pp_cost))
        monthly_sales_r = 0;   monthly_sales_l = 0
        month_end_cost.append(monthly_p_cost + monthly_fixed_cost)
        monthly_p_cost = 0 
        profit = np.subtract(month_end_Tsales, month_end_cost )
        #print(profit)

## ------------------------------------------------------------------------ ##

## ------------------------------------------------------------------------ ##
#print(month_end_sales_r)

sales_doc = load_workbook('Data/weekly_sales.xlsx')
sheet = sales_doc.active
sheet.cell(row=1, column=6).value = "result_r"
sheet.cell(row=1,column=7).value = "result_l"
j=2
for i in range (0, len(result_r)):
    sheet.cell(row=j,column=6).value = result_r[i]
    sheet.cell(row=j,column=7).value = result_l[i]
    j += 1

sheet.cell(row=1,column=8).value = "monthly_end_cost"
sheet.cell(row=1,column=9).value = "month_end_Tsales"
sheet.cell(row=1,column=10).value = "profit"
k = 5
for i in range(0, len(month_end_cost)):
    sheet.cell(row=k, column=8).value = month_end_cost[i]
    sheet.cell(row=k, column=9).value = month_end_Tsales[i]
    sheet.cell(row=k, column=10).value = profit[i]
    k += 4
    
sales_doc.save("Data/m_weekly_sales.xlsx")

